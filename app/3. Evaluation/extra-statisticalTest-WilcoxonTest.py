import openpyxl
import numpy as np
from scipy.stats import wilcoxon, rankdata

VRAG_FILE  = "3_1_vRAG_evaluationResult.xlsx"
KGRAG_FILE = "graphRAG_evaluationResult_20260715_130836.xlsx"

def answer_accuracy_per_question(path):
    """Average per-question Answer Accuracy over the three judge runs."""
    wb = openpyxl.load_workbook(path, data_only=True)
    runs = []
    for sheet in ["4omini_run1", "4omini_run2", "4omini_run3"]:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        col = header.index("Answer Accuracy")
        runs.append([
            (r[col] if isinstance(r[col], (int, float)) else np.nan)
            for r in ws.iter_rows(min_row=2, values_only=True)
        ])
    return np.nanmean(np.array(runs, dtype=float), axis=0)   # length 35

vrag  = answer_accuracy_per_question(VRAG_FILE)
kgrag = answer_accuracy_per_question(KGRAG_FILE)

print(f"VRAG  mean Answer Accuracy: {np.nanmean(vrag):.3f}")
print(f"KGRAG mean Answer Accuracy: {np.nanmean(kgrag):.3f}")

# paired Wilcoxon signed-rank test (all 35 questions)
diff = kgrag - vrag
w_stat, p_two = wilcoxon(kgrag, vrag, zero_method="wilcox", alternative="two-sided")

# matched-pairs rank-biserial effect size
nz = diff[diff != 0]
ranks = rankdata(np.abs(nz))
r_plus, r_minus = ranks[nz > 0].sum(), ranks[nz < 0].sum()
rank_biserial = (r_plus - r_minus) / (r_plus + r_minus)

print(f"\nKGRAG better: {int((diff > 0).sum())} | "
      f"VRAG better: {int((diff < 0).sum())} | ties: {int((diff == 0).sum())}")
print(f"W = {w_stat:.1f}")
print(f"p (two-sided) = {p_two:.4f}")
print(f"rank-biserial effect size r = {rank_biserial:.3f}")e